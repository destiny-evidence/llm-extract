import json
import typing

import dspy
import openpyxl
from pydantic import Field
from pydantic.dataclasses import dataclass as pydantic_dataclass

from llm_extract.export import (
    ExtractionResult,
    NOT_FOUND,
    format_value,
    write_extraction_results_to_folder,
)
from llm_extract.models import Attribute


def test_format_value_none_returns_not_found() -> None:
    assert format_value(None) == NOT_FOUND


def test_format_value_string_strips_double_quotes() -> None:
    assert format_value('"Aeron Chair"') == "Aeron Chair"


def test_format_value_string_strips_single_quotes() -> None:
    assert format_value("'Aeron Chair'") == "Aeron Chair"


def test_format_value_plain_string_unchanged() -> None:
    assert format_value("Aeron Chair") == "Aeron Chair"


def test_format_value_list_json_serialised() -> None:
    assert format_value([1, 2, 3]) == "[1, 2, 3]"


def test_format_value_dict_json_serialised() -> None:
    assert format_value({"width": 68.5}) == '{"width": 68.5}'


def test_format_value_primitives_unchanged() -> None:
    assert format_value(42) == 42
    assert format_value(14.99) == 14.99
    assert format_value(True) is True


@pydantic_dataclass
class _InterventionType:
    type_of_intervention: typing.Optional[str] = Field(default=None)


@pydantic_dataclass
class _Intervention:
    group_name: typing.Optional[str] = Field(default=None)
    intervention_type: typing.Optional[_InterventionType] = Field(default=None)


def test_format_value_list_of_nested_dataclasses_json_serialised() -> None:
    value = [
        _Intervention(group_name="Risperidone", intervention_type=None),
        _Intervention(group_name="Haloperidol", intervention_type=None),
    ]
    assert format_value(value) == (
        '[{"group_name": "Risperidone", "intervention_type": null}, '
        '{"group_name": "Haloperidol", "intervention_type": null}]'
    )


def test_to_csv_rows_has_header() -> None:
    result = ExtractionResult(prediction=dspy.Prediction(product_name="Widget"))
    assert result.to_csv_rows()[0] == ["name", "value"]


def test_to_csv_rows_contains_extracted_values() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", price=14.99)
    )
    rows = result.to_csv_rows()
    assert ["product_name", "Widget"] in rows
    assert ["price", 14.99] in rows


def test_to_csv_rows_none_value_becomes_not_found() -> None:
    result = ExtractionResult(prediction=dspy.Prediction(product_name=None))
    assert ["product_name", NOT_FOUND] in result.to_csv_rows()


def test_to_csv_rows_no_reasoning_row_when_absent() -> None:
    result = ExtractionResult(prediction=dspy.Prediction(product_name="Widget"))
    names = [row[0] for row in result.to_csv_rows()]
    assert "_reasoning_" not in names


def test_to_csv_rows_reasoning_appended_as_last_row() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", reasoning="Found in line 1.")
    )
    assert result.to_csv_rows()[-1] == ["_reasoning_", "Found in line 1."]


def test_write_csv_creates_file(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", price=14.99)
    )
    output = tmp_path / "results.csv"
    result.write_csv(output)
    assert output.exists()


def test_write_csv_content(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", price=14.99)
    )
    output = tmp_path / "results.csv"
    result.write_csv(output)
    content = output.read_text()
    assert "product_name" in content
    assert "Widget" in content
    assert "price" in content
    assert "14.99" in content


def test_to_csv_rows_reasoning_excluded_from_regular_rows() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", reasoning="Found in line 1.")
    )
    regular_names = [row[0] for row in result.to_csv_rows()[1:-1]]
    assert "reasoning" not in regular_names


# --- write_excel ---


@pydantic_dataclass
class _OutcomeCategory:
    value: typing.Optional[str] = Field(default=None)


@pydantic_dataclass
class _Outcome:
    name: typing.Optional[str] = Field(default=None)
    category: typing.Optional[_OutcomeCategory] = Field(default=None)
    sub_interventions: typing.Optional[list[_Intervention]] = Field(default=None)


def test_write_excel_creates_file(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", price=14.99),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
            Attribute(name="price", attr_type=typing.Optional[float], description=""),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)
    assert output.exists()


def test_write_excel_plain_attrs_in_summary_sheet(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", price=14.99),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
            Attribute(name="price", attr_type=typing.Optional[float], description=""),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    assert "Summary" in workbook.sheetnames
    rows = list(workbook["Summary"].iter_rows(values_only=True))
    assert rows[0] == ("name", "value")
    assert ("product_name", "Widget") in rows
    assert ("price", 14.99) in rows


def test_write_excel_reasoning_in_summary_sheet(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", reasoning="Found in line 1."),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    rows = list(workbook["Summary"].iter_rows(values_only=True))
    assert ("_reasoning_", "Found in line 1.") in rows


def test_write_excel_list_of_custom_type_gets_own_sheet(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            interventions=[
                _Intervention(
                    group_name="Risperidone",
                    intervention_type=_InterventionType("Intervention"),
                ),
                _Intervention(group_name="Haloperidol", intervention_type=None),
            ]
        ),
        attributes=[
            Attribute(
                name="interventions",
                attr_type=typing.Optional[list[_Intervention]],
                description="",
            ),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    assert "interventions" in workbook.sheetnames
    rows = list(workbook["interventions"].iter_rows(values_only=True))
    assert rows[0] == ("group_name", "intervention_type.type_of_intervention")
    assert rows[1] == ("Risperidone", "Intervention")
    assert rows[2] == ("Haloperidol", NOT_FOUND)

    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    assert ("interventions", "interventions") in summary_rows
    link_cell = workbook["Summary"]["B2"]
    assert link_cell.hyperlink.target == "#'interventions'!A1"


def test_write_excel_scalar_custom_type_gets_single_row_sheet(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            lead_intervention=_Intervention(
                group_name="Risperidone",
                intervention_type=_InterventionType("Intervention"),
            )
        ),
        attributes=[
            Attribute(
                name="lead_intervention",
                attr_type=typing.Optional[_Intervention],
                description="",
            ),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    assert "lead_intervention" in workbook.sheetnames
    rows = list(workbook["lead_intervention"].iter_rows(values_only=True))
    assert rows[0] == ("group_name", "intervention_type.type_of_intervention")
    assert rows[1] == ("Risperidone", "Intervention")

    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    assert ("lead_intervention", "lead_intervention") in summary_rows


def test_write_excel_missing_custom_type_has_not_found_in_summary_and_no_sheet(
    tmp_path,
) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(interventions=None),
        attributes=[
            Attribute(
                name="interventions",
                attr_type=typing.Optional[list[_Intervention]],
                description="",
            ),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    assert "interventions" not in workbook.sheetnames
    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    assert ("interventions", NOT_FOUND) in summary_rows


def test_write_excel_empty_list_custom_type_has_not_found_in_summary_and_no_sheet(
    tmp_path,
) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(interventions=[]),
        attributes=[
            Attribute(
                name="interventions",
                attr_type=typing.Optional[list[_Intervention]],
                description="",
            ),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    assert "interventions" not in workbook.sheetnames
    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    assert ("interventions", NOT_FOUND) in summary_rows


def test_write_excel_deeply_nested_list_gets_own_pooled_sheet(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            outcomes=[
                _Outcome(
                    name="BPRS score",
                    category=_OutcomeCategory("Mental Health Outcome"),
                    sub_interventions=[
                        _Intervention(group_name="Risperidone", intervention_type=None)
                    ],
                )
            ]
        ),
        attributes=[
            Attribute(
                name="outcomes",
                attr_type=typing.Optional[list[_Outcome]],
                description="",
            ),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    rows = list(workbook["outcomes"].iter_rows(values_only=True))
    assert rows[0] == ("name", "category.value", "sub_interventions")
    assert rows[1][0:2] == ("BPRS score", "Mental Health Outcome")
    assert rows[1][2] == "_Intervention"

    link_cell = workbook["outcomes"]["C2"]
    assert link_cell.hyperlink.target == "#'_Intervention'!A1"

    intervention_rows = list(workbook["_Intervention"].iter_rows(values_only=True))
    assert intervention_rows[0] == (
        "_parent_sheet",
        "_parent_row",
        "group_name",
        "intervention_type.type_of_intervention",
    )
    assert intervention_rows[1] == ("outcomes", 1, "Risperidone", NOT_FOUND)


def test_write_excel_pooled_sheet_links_multiple_parent_rows(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            outcomes=[
                _Outcome(
                    name="BPRS score",
                    category=None,
                    sub_interventions=[
                        _Intervention(group_name="Risperidone", intervention_type=None)
                    ],
                ),
                _Outcome(
                    name="CGI score",
                    category=None,
                    sub_interventions=[
                        _Intervention(group_name="Haloperidol", intervention_type=None),
                        _Intervention(group_name="Placebo", intervention_type=None),
                    ],
                ),
            ]
        ),
        attributes=[
            Attribute(
                name="outcomes",
                attr_type=typing.Optional[list[_Outcome]],
                description="",
            ),
        ],
    )
    output = tmp_path / "results.xlsx"
    result.write_excel(output)

    workbook = openpyxl.load_workbook(output)
    intervention_rows = list(workbook["_Intervention"].iter_rows(values_only=True))
    assert intervention_rows[1] == ("outcomes", 1, "Risperidone", NOT_FOUND)
    assert intervention_rows[2] == ("outcomes", 2, "Haloperidol", NOT_FOUND)
    assert intervention_rows[3] == ("outcomes", 2, "Placebo", NOT_FOUND)


# --- to_json / write_json ---


def test_to_json_plain_attrs() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", price=14.99),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
            Attribute(name="price", attr_type=typing.Optional[float], description=""),
        ],
    )
    assert result.to_json() == {"product_name": "Widget", "price": 14.99}


def test_to_json_missing_value_is_not_found() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name=None),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
        ],
    )
    assert result.to_json() == {"product_name": NOT_FOUND}


def test_to_json_no_reasoning_key_when_absent() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget"),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
        ],
    )
    assert "_reasoning_" not in result.to_json()


def test_to_json_reasoning_included_as_underscored_key() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", reasoning="Found in line 1."),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
        ],
    )
    assert result.to_json()["_reasoning_"] == "Found in line 1."


def test_to_json_without_attributes_falls_back_to_prediction_keys() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget", price=14.99)
    )
    assert result.to_json() == {"product_name": "Widget", "price": 14.99}


def test_to_json_nested_custom_type_preserved_as_plain_dict() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            lead_intervention=_Intervention(
                group_name="Risperidone",
                intervention_type=_InterventionType("Intervention"),
            )
        ),
        attributes=[
            Attribute(
                name="lead_intervention",
                attr_type=typing.Optional[_Intervention],
                description="",
            ),
        ],
    )
    assert result.to_json() == {
        "lead_intervention": {
            "group_name": "Risperidone",
            "intervention_type": {"type_of_intervention": "Intervention"},
        }
    }


def test_to_json_list_of_custom_type_preserved_as_plain_list() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            interventions=[
                _Intervention(group_name="Risperidone", intervention_type=None),
                _Intervention(group_name="Haloperidol", intervention_type=None),
            ]
        ),
        attributes=[
            Attribute(
                name="interventions",
                attr_type=typing.Optional[list[_Intervention]],
                description="",
            ),
        ],
    )
    assert result.to_json() == {
        "interventions": [
            {"group_name": "Risperidone", "intervention_type": NOT_FOUND},
            {"group_name": "Haloperidol", "intervention_type": NOT_FOUND},
        ]
    }


def test_to_json_deeply_nested_list_preserved_without_flattening() -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            outcomes=[
                _Outcome(
                    name="BPRS score",
                    category=_OutcomeCategory("Mental Health Outcome"),
                    sub_interventions=[
                        _Intervention(group_name="Risperidone", intervention_type=None)
                    ],
                )
            ]
        ),
        attributes=[
            Attribute(
                name="outcomes",
                attr_type=typing.Optional[list[_Outcome]],
                description="",
            ),
        ],
    )
    assert result.to_json() == {
        "outcomes": [
            {
                "name": "BPRS score",
                "category": {"value": "Mental Health Outcome"},
                "sub_interventions": [
                    {"group_name": "Risperidone", "intervention_type": NOT_FOUND}
                ],
            }
        ]
    }


def test_write_json_creates_file(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget"),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
        ],
    )
    output = tmp_path / "results.json"
    result.write_json(output)
    assert output.exists()


def test_write_json_content_round_trips_via_json_load(tmp_path) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(
            lead_intervention=_Intervention(
                group_name="Risperidone", intervention_type=None
            )
        ),
        attributes=[
            Attribute(
                name="lead_intervention",
                attr_type=typing.Optional[_Intervention],
                description="",
            ),
        ],
    )
    output = tmp_path / "results.json"
    result.write_json(output)

    with output.open() as f:
        loaded = json.load(f)
    assert loaded == {
        "lead_intervention": {
            "group_name": "Risperidone",
            "intervention_type": NOT_FOUND,
        }
    }


# --- write_extraction_results_to_folder ---


def test_write_extraction_results_to_folder_json_not_written_by_default(
    tmp_path,
) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget"),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
        ],
    )
    output_dir = tmp_path / "out"
    write_extraction_results_to_folder(output_dir, [("file1", result)])

    assert (output_dir / "file1-extracted.csv").exists()
    assert not (output_dir / "file1-extracted.json").exists()


def test_write_extraction_results_to_folder_also_json_writes_both_formats(
    tmp_path,
) -> None:
    result = ExtractionResult(
        prediction=dspy.Prediction(product_name="Widget"),
        attributes=[
            Attribute(
                name="product_name", attr_type=typing.Optional[str], description=""
            ),
        ],
    )
    output_dir = tmp_path / "out"
    write_extraction_results_to_folder(
        output_dir, [("file1", result)], use_excel=True, also_json=True
    )

    assert (output_dir / "file1-extracted.xlsx").exists()
    assert (output_dir / "file1-extracted.json").exists()

    with (output_dir / "file1-extracted.json").open() as f:
        loaded = json.load(f)
    assert loaded == {"product_name": "Widget"}
