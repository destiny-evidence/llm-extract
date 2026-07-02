import openpyxl
import pytest
from pathlib import Path
from llm_extract.loader import load_csv, load_workbook
from llm_extract.models import CSVData, WorkbookData


def test_load_csv_valid(tmp_path: Path) -> None:
    csv = tmp_path / "attrs.csv"
    csv.write_text(
        "name,type,description\nproduct_name,str,The product name\nprice,float,The price\n"
    )
    data = load_csv(csv)
    assert isinstance(data, CSVData)
    assert len(data.rows) == 2
    assert data.rows[0]["name"] == "product_name"
    assert data.rows[1]["name"] == "price"


def test_load_csv_empty_returns_empty_list(tmp_path: Path) -> None:
    csv = tmp_path / "attrs.csv"
    csv.write_text("name,type,description\n")
    data = load_csv(csv)
    assert data.rows == []


def test_load_csv_missing_column_raises(tmp_path: Path) -> None:
    csv = tmp_path / "attrs.csv"
    csv.write_text("name,type\nfoo,str\n")
    with pytest.raises(ValueError, match="CSV missing columns"):
        load_csv(csv)


def test_load_csv_missing_all_required_columns_raises(tmp_path: Path) -> None:
    csv = tmp_path / "attrs.csv"
    csv.write_text("col_a,col_b\nfoo,bar\n")
    with pytest.raises(ValueError, match="CSV missing columns"):
        load_csv(csv)


def test_load_csv_accepts_string_path(tmp_path: Path) -> None:
    csv = tmp_path / "attrs.csv"
    csv.write_text("name,type,description\nfoo,str,Foo\n")
    data1 = load_csv(csv)
    data2 = load_csv(str(csv))
    assert data1.rows == data2.rows


# --- load_workbook ---


def _write_workbook(path: Path, sheets: dict[str, list[list]]) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def test_load_workbook_valid(tmp_path: Path) -> None:
    path = tmp_path / "types.xlsx"
    _write_workbook(
        path,
        {
            "Study": [
                ["name", "type", "description"],
                ["title", "str", "The title"],
            ]
        },
    )
    data = load_workbook(path)
    assert isinstance(data, WorkbookData)
    assert data.sheets == {
        "Study": [{"name": "title", "type": "str", "description": "The title"}]
    }


def test_load_workbook_multiple_sheets(tmp_path: Path) -> None:
    path = tmp_path / "types.xlsx"
    _write_workbook(
        path,
        {
            "Study": [
                ["name", "type", "description"],
                ["title", "str", "The title"],
            ],
            "Author": [
                ["name", "type", "description"],
                ["full_name", "str", "The author's name"],
            ],
        },
    )
    data = load_workbook(path)
    assert set(data.sheets) == {"Study", "Author"}
    assert data.sheets["Author"] == [
        {"name": "full_name", "type": "str", "description": "The author's name"}
    ]


def test_load_workbook_empty_sheet_returns_empty_list(tmp_path: Path) -> None:
    path = tmp_path / "types.xlsx"
    _write_workbook(path, {"Study": []})
    data = load_workbook(path)
    assert data.sheets == {"Study": []}


def test_load_workbook_skips_blank_rows(tmp_path: Path) -> None:
    path = tmp_path / "types.xlsx"
    _write_workbook(
        path,
        {
            "Study": [
                ["name", "type", "description"],
                ["title", "str", "The title"],
                [None, None, None],
            ]
        },
    )
    data = load_workbook(path)
    assert data.sheets == {
        "Study": [{"name": "title", "type": "str", "description": "The title"}]
    }


def test_load_workbook_missing_columns_raises(tmp_path: Path) -> None:
    path = tmp_path / "types.xlsx"
    _write_workbook(path, {"Study": [["name", "type"], ["title", "str"]]})
    with pytest.raises(ValueError, match="missing columns"):
        load_workbook(path)


def test_load_workbook_accepts_string_path(tmp_path: Path) -> None:
    path = tmp_path / "types.xlsx"
    _write_workbook(
        path,
        {
            "Study": [
                ["name", "type", "description"],
                ["title", "str", "The title"],
            ]
        },
    )
    data1 = load_workbook(path)
    data2 = load_workbook(str(path))
    assert data1.sheets == data2.sheets
