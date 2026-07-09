import csv
import dataclasses
import json
import re
import typing
from pathlib import Path
from typing import Callable

import dspy
import openpyxl
from pydantic_core import to_jsonable_python

from llm_extract.models import Attribute

NOT_FOUND = "NOT_FOUND"

_INVALID_SHEET_NAME_CHARS = re.compile(r"[:\\/?*\[\]]")


def _format_value(value: object) -> object:
    if value is None:
        return NOT_FOUND
    if isinstance(value, str):
        return value.strip("\"'")
    if isinstance(value, (list, dict)):
        # e.g. a list[type] value from a plain CSV attrs file - to_jsonable_python
        # handles any nesting consistently before json.dumps.
        return json.dumps(to_jsonable_python(value))
    return value


def _apply_not_found_sentinel(value: object) -> object:
    """
    Recursively replace `None` with the NOT_FOUND sentinel, matching the
    missing-value convention CSV/Excel already apply via `_format_value`.

    :param value: a plain JSON-able value (e.g. from `to_jsonable_python`),
                   possibly containing `None` at any nesting depth
    :return: the same structure with every `None` replaced by NOT_FOUND
    """
    if value is None:
        return NOT_FOUND
    if isinstance(value, list):
        return [_apply_not_found_sentinel(v) for v in value]
    if isinstance(value, dict):
        return {k: _apply_not_found_sentinel(v) for k, v in value.items()}
    return value


def _unwrap_optional(type_: object) -> object:
    """
    Unwrap an `Optional[X]` annotation to `X`.

    :param type_: a type annotation, possibly `Optional[X]`
    :return: `X` if `type_` is `Optional[X]`, otherwise `type_` unchanged
    """
    args = typing.get_args(type_)
    if type(None) in args:
        non_none = [arg for arg in args if arg is not type(None)]
        if len(non_none) == 1:
            return non_none[0]
    return type_


def _sanitize_sheet_name(name: str) -> str:
    """
    Make a string safe to use as an Excel sheet title.

    :param name: the proposed sheet name
    :return: `name` with characters invalid in Excel sheet titles replaced
             by underscores and truncated to Excel's 31 character limit
    """
    return _INVALID_SHEET_NAME_CHARS.sub("_", name)[:31]


def _flatten_columns(dataclass_type: type, prefix: str = "") -> list[str]:
    """
    Compute flattened column names for a dataclass type.

    Fields whose type is itself a dataclass are flattened into dot-prefixed
    columns (e.g. "intervention_type.type_of_intervention"). Any other
    field, including a list of dataclasses, becomes a single column.

    :param dataclass_type: the dataclass type to flatten
    :param prefix: prefix to prepend to each column name
    :return: list of flattened column names
    """
    columns = []
    for field in dataclasses.fields(dataclass_type):
        inner_type = _unwrap_optional(field.type)
        name = f"{prefix}{field.name}"
        if dataclasses.is_dataclass(inner_type):
            columns.extend(_flatten_columns(inner_type, prefix=f"{name}."))
        else:
            columns.append(name)
    return columns


def _flatten_instance(instance: object, dataclass_type: type, prefix: str = "") -> dict:
    """
    Flatten a dataclass instance into a dict of column name to formatted value.

    :param instance: the dataclass instance to flatten, or None
    :param dataclass_type: the dataclass type describing `instance`'s shape
    :param prefix: prefix to prepend to each column name
    :return: dict mapping flattened column names to values formatted via
             :func:`_format_value`
    """
    row = {}
    for field in dataclasses.fields(dataclass_type):
        inner_type = _unwrap_optional(field.type)
        name = f"{prefix}{field.name}"
        value = getattr(instance, field.name, None) if instance is not None else None
        if dataclasses.is_dataclass(inner_type):
            row.update(_flatten_instance(value, inner_type, prefix=f"{name}."))
        else:
            row[name] = _format_value(value)
    return row


def _flatten_list_fields(
    instance: object, dataclass_type: type, prefix: str = ""
) -> list[tuple[str, type, list]]:
    """
    Find dot-prefixed columns (matching :func:`_flatten_columns`' naming) whose field
    type is `list[dataclass]`, recursing through nested singular dataclasses the same
    way `_flatten_columns` does so the two stay in sync.

    A `list[dataclass]` field can't be inlined as fixed columns (arbitrary length), so
    instead of JSON-encoding it into one cell, callers use this to spawn a pooled child
    table for `elem_type` and link to it.

    :param instance: the dataclass instance to inspect, or None
    :param dataclass_type: the dataclass type describing `instance`'s shape
    :param prefix: prefix to prepend to each column name
    :return: list of (column_name, element dataclass type, list of item instances)
    """
    results = []
    for field in dataclasses.fields(dataclass_type):
        inner_type = _unwrap_optional(field.type)
        name = f"{prefix}{field.name}"
        value = getattr(instance, field.name, None) if instance is not None else None
        if dataclasses.is_dataclass(inner_type):
            results.extend(_flatten_list_fields(value, inner_type, prefix=f"{name}."))
        elif typing.get_origin(inner_type) is list:
            (elem_type,) = typing.get_args(inner_type)
            if dataclasses.is_dataclass(elem_type):
                results.append((name, elem_type, value or []))
    return results


@dataclasses.dataclass(frozen=True)
class _SheetLink:
    """Placeholder cell value meaning "link to this other sheet" rather than raw data."""

    sheet_name: str


@dataclasses.dataclass
class _Table:
    """
    Flattened tabular data for a custom-type attribute or a pooled list[dataclass] field.

    `include_parent_columns` distinguishes root attribute tables (exactly one implicit
    parent - the Summary row, so a parent link adds no information) from pooled child
    tables (reached from a list[dataclass] field at any depth, which can have rows from
    multiple distinct parent rows, so each row needs its own parent link).
    """

    columns: list[str]
    rows: list[list]
    parent_refs: list[tuple[str, int]] = dataclasses.field(default_factory=list)
    include_parent_columns: bool = False


def _process_record(
    instance: object,
    dataclass_type: type,
    sheet_name: str,
    parent_ref: tuple[str, int] | None,
    tables: dict[str, _Table],
    include_parent_columns: bool,
) -> int:
    """
    Flatten `instance` into a new row of `tables[sheet_name]`, creating the table on
    first use, and recurse into any nested list[dataclass] fields (at any depth) to
    populate their own pooled child tables instead of JSON-encoding them into one cell.

    :param instance: the dataclass instance to flatten into a row
    :param dataclass_type: the dataclass type describing `instance`'s shape
    :param sheet_name: name of the table this row belongs to
    :param parent_ref: (parent_sheet_name, parent_row_number) this row is linked from,
                        or None for a root attribute's table
    :param tables: shared dict of sheet name to `_Table`, populated in place
    :param include_parent_columns: whether this row should carry a parent link
    :return: 1-indexed row number of the newly appended row within `tables[sheet_name]`
    """
    columns = _flatten_columns(dataclass_type)
    values = _flatten_instance(instance, dataclass_type)
    list_fields = _flatten_list_fields(instance, dataclass_type)

    table = tables.setdefault(
        sheet_name,
        _Table(columns=columns, rows=[], include_parent_columns=include_parent_columns),
    )
    row_values = [values[column] for column in columns]
    row_number = len(table.rows) + 1

    for column_name, elem_type, sub_items in list_fields:
        col_index = columns.index(column_name)
        if not sub_items:
            row_values[col_index] = NOT_FOUND
            continue
        child_sheet = _sanitize_sheet_name(elem_type.__name__)
        row_values[col_index] = _SheetLink(child_sheet)
        for sub_item in sub_items:
            _process_record(
                sub_item,
                elem_type,
                child_sheet,
                (sheet_name, row_number),
                tables,
                include_parent_columns=True,
            )

    table.rows.append(row_values)
    if include_parent_columns:
        table.parent_refs.append(parent_ref)
    return row_number


def _classify_attribute(
    attr: Attribute, prediction: dspy.Prediction, tables: dict[str, _Table]
) -> object:
    """
    Classify a top-level attribute's value for presentation.

    :param attr: the attribute definition, used for its type
    :param prediction: the DSPy Prediction holding the extracted value
    :param tables: shared dict of sheet name to `_Table`, populated in place for
                    custom-type attributes
    :return: a `_SheetLink` if `attr` is a custom type (or a list of one) with a
             non-empty value, NOT_FOUND if it's a custom type with no value, otherwise
             the value formatted via :func:`_format_value`
    """
    value = getattr(prediction, attr.name, None)
    inner_type = _unwrap_optional(attr.attr_type)

    dataclass_type = None
    items = []
    if typing.get_origin(inner_type) is list:
        (elem_type,) = typing.get_args(inner_type)
        if dataclasses.is_dataclass(elem_type):
            dataclass_type, items = elem_type, value or []
    elif dataclasses.is_dataclass(inner_type):
        dataclass_type = inner_type
        items = [value] if value is not None else []

    if dataclass_type is None:
        return _format_value(value)
    if not items:
        return NOT_FOUND

    sheet_name = _sanitize_sheet_name(attr.name)
    for item in items:
        _process_record(
            item, dataclass_type, sheet_name, None, tables, include_parent_columns=False
        )
    return _SheetLink(sheet_name)


def _append_link_row(
    sheet: openpyxl.worksheet.worksheet.Worksheet, name: str, target_sheet_name: str
) -> None:
    """
    Append a [name, link] row, where the link cell jumps to another sheet.

    :param sheet: the sheet to append the row to
    :param name: value for the first ("name") column
    :param target_sheet_name: title of the sheet the link cell should jump to
    """
    sheet.append([name, target_sheet_name])
    link_cell = sheet.cell(row=sheet.max_row, column=2)
    link_cell.hyperlink = f"#'{target_sheet_name}'!A1"
    link_cell.style = "Hyperlink"


@dataclasses.dataclass
class ExtractionResult:
    """Wraps a DSPy Prediction with CSV and Excel serialisation."""

    prediction: dspy.Prediction
    attributes: list[Attribute] = dataclasses.field(default_factory=list)

    def to_csv_rows(self) -> list[list]:
        """
        Serialise extraction results to a list of rows ready for csv.writer.

        :return: list of [name, value] rows, with a header row first and an
                 optional [_reasoning_, ...] row last if reasoning was produced
        """
        rows = [["name", "value"]]

        for key in (k for k in self.prediction.keys() if k != "reasoning"):
            value = getattr(self.prediction, key)
            rows.append([key, _format_value(value)])

        reasoning = getattr(self.prediction, "reasoning", None)
        if reasoning is not None:
            rows.append(["_reasoning_", reasoning])

        return rows

    def write_csv(self, path: Path) -> None:
        """
        Write extraction results to a CSV file.

        :param path: destination file path
        """
        with path.open("w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(self.to_csv_rows())

    def to_json(self) -> dict:
        """
        Serialise extraction results to a plain JSON-able dict, preserving the
        full nested structure (no flattening, no JSON-in-a-cell).

        :return: dict mapping each attribute name to its value, converted via
                 `to_jsonable_python` (handles our pydantic-dataclass-based
                 custom types, which `Prediction.toDict()` does not - it only
                 recognises genuine `pydantic.BaseModel` instances). Missing
                 values are the NOT_FOUND sentinel, matching CSV/Excel, applied
                 at any nesting depth (e.g. a null field inside a custom type,
                 or a missing custom-type attribute). Includes a "_reasoning_"
                 key if reasoning was produced.
        """
        keys = [attr.name for attr in self.attributes] or [
            k for k in self.prediction.keys() if k != "reasoning"
        ]
        result = {
            key: _apply_not_found_sentinel(
                to_jsonable_python(getattr(self.prediction, key, None))
            )
            for key in keys
        }

        reasoning = getattr(self.prediction, "reasoning", None)
        if reasoning is not None:
            result["_reasoning_"] = reasoning

        return result

    def write_json(self, path: Path) -> None:
        """
        Write extraction results to a JSON file, preserving the full nested
        structure so results can be consumed programmatically without
        parsing CSV/Excel.

        :param path: destination file path
        """
        with path.open("w") as f:
            json.dump(self.to_json(), f, indent=2)

    def write_excel(self, path: Path) -> None:
        """
        Write extraction results to an Excel workbook.

        A "Summary" sheet has one row per top-level attribute. Plain
        attributes show their formatted value directly. Attributes that are
        a custom type, or a non-empty list of a custom type, instead get a
        link to a dedicated sheet named after the attribute: one row per
        item, with columns for each field. Fields that are themselves a
        custom type are flattened into dot-prefixed columns (e.g.
        "intervention_type.type_of_intervention"), at any nesting depth.
        Fields that are a *list* of a custom type - at any depth, not just
        top-level attributes - instead link to their own sheet named after
        that type, pooling every instance of it found anywhere in the
        results. Each row in a pooled sheet carries "_parent_sheet" and
        "_parent_row" columns identifying the specific row that produced it,
        since a pooled type can be reached from more than one place.
        Custom-type attributes with no value are shown as NOT_FOUND in the
        Summary sheet, with no dedicated sheet.

        :param path: destination file path
        """
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Summary")
        summary.append(["name", "value"])

        tables: dict[str, _Table] = {}
        for attr in self.attributes:
            result = _classify_attribute(attr, self.prediction, tables)
            if isinstance(result, _SheetLink):
                _append_link_row(summary, attr.name, result.sheet_name)
            else:
                summary.append([attr.name, result])

        reasoning = getattr(self.prediction, "reasoning", None)
        if reasoning is not None:
            summary.append(["_reasoning_", reasoning])

        for sheet_name, table in tables.items():
            sheet = workbook.create_sheet(sheet_name)
            header = (
                ["_parent_sheet", "_parent_row"] + table.columns
                if table.include_parent_columns
                else table.columns
            )
            sheet.append(header)

            parent_refs = (
                table.parent_refs
                if table.include_parent_columns
                else [None] * len(table.rows)
            )
            link_offset = 2 if table.include_parent_columns else 0
            for row_values, parent_ref in zip(table.rows, parent_refs):
                display_values = [
                    v.sheet_name if isinstance(v, _SheetLink) else v for v in row_values
                ]
                if table.include_parent_columns:
                    display_values = [parent_ref[0], parent_ref[1]] + display_values
                sheet.append(display_values)

                for col_index, value in enumerate(row_values):
                    if isinstance(value, _SheetLink):
                        cell = sheet.cell(
                            row=sheet.max_row, column=col_index + 1 + link_offset
                        )
                        cell.hyperlink = f"#'{value.sheet_name}'!A1"
                        cell.style = "Hyperlink"

        workbook.save(path)


def write_extraction_results_to_folder(
    output_dir: Path,
    results: list[tuple[str, ExtractionResult]],
    use_excel: bool = False,
    also_json: bool = False,
    on_progress: Callable[[int, int], None] | None = None,
) -> None:
    """
    Write multiple extraction results to a folder with one file per result.

    Each file is named <filename>-extracted.csv or .xlsx depending on use_excel.
    Preserves directory structure when results include relative paths with subdirectories.

    :param output_dir: directory to write results to (created if it doesn't exist)
    :param results: list of (filename_or_relative_path, ExtractionResult) tuples
    :param use_excel: whether to write Excel files (True) or CSV files (False)
    :param also_json: whether to additionally write a <filename>-extracted.json
                       file alongside the csv/xlsx file, for programmatic consumption
    :param on_progress: optional callback (current, total) called after each file is written
    :return: None
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    for i, (filename, result) in enumerate(results):
        extension = "xlsx" if use_excel else "csv"
        file_path = output_dir / f"{filename}-extracted.{extension}"
        file_path.parent.mkdir(parents=True, exist_ok=True)
        if use_excel:
            result.write_excel(file_path)
        else:
            result.write_csv(file_path)

        if also_json:
            result.write_json(output_dir / f"{filename}-extracted.json")

        if on_progress:
            on_progress(i + 1, len(results))
