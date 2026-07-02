import csv
from pathlib import Path

import openpyxl

from llm_extract.models import CSVData, WorkbookData
from llm_extract.exceptions import LoadingAttributeFromCSVError

EXPECTED_COLUMNS = {"name", "type", "description"}


def load_csv(path: Path | str) -> CSVData:
    """
    Load attribute data from a CSV file.

    :param path: path to the CSV file with columns: name, type, description
    :return: CSVData containing the loaded rows
    :raises ValueError: if required columns are missing
    :raises LoadingAttributeFromCSVError: if the file cannot be read
    """
    path = Path(path)
    try:
        with path.open() as f:
            reader = csv.DictReader(f)
            missing = EXPECTED_COLUMNS - set(reader.fieldnames or [])
            if missing:
                raise ValueError(f"CSV missing columns: {missing}")
            rows = list(reader)
        return CSVData(rows=rows)
    except ValueError:
        raise
    except Exception as exc:
        raise LoadingAttributeFromCSVError(f"Failed to load CSV: {exc}")


def load_workbook(path: Path | str) -> WorkbookData:
    """
    Load all sheets from an Excel workbook as raw attribute rows.

    Each sheet represents a user-defined custom type, and must have 'name',
    'type' and 'description' columns, with each row describing one field of
    that type.

    :param path: path to the Excel workbook
    :return: WorkbookData containing all sheets
    :raises ValueError: if any sheet is missing required columns
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in workbook.sheetnames:
        rows = workbook[name].iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            sheets[name] = []
            continue
        missing = EXPECTED_COLUMNS - set(header)
        if missing:
            raise ValueError(f"Sheet '{name}' missing columns: {missing}")
        sheets[name] = [
            dict(zip(header, row))
            for row in rows
            if any(cell is not None for cell in row)
        ]
    return WorkbookData(sheets=sheets)
