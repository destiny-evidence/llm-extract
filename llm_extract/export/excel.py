import dataclasses
import re
import typing
from pathlib import Path

import dspy
import openpyxl

from llm_extract.export.common import NOT_FOUND, format_value, unwrap_optional
from llm_extract.models import Attribute

_INVALID_SHEET_NAME_CHARS = re.compile(r"[:\\/?*\[\]]")


def _sanitize_sheet_name(name: str) -> str:
    """
    Make a string safe to use as an Excel sheet title.

    :param name: the proposed sheet name
    :return: `name` with characters invalid in Excel sheet titles replaced
             by underscores and truncated to Excel's 31 character limit
    """
    return _INVALID_SHEET_NAME_CHARS.sub("_", name)[:31]


def _flatten_columns(dataclass_type: type, prefix: str = "") -> list[str]:
    """
    Compute flattened column names for a dataclass type.

    Fields whose type is itself a dataclass are flattened into dot-prefixed
    columns (e.g. "intervention_type.type_of_intervention"). Any other
    field, including a list of dataclasses, becomes a single column.

    :param dataclass_type: the dataclass type to flatten
    :param prefix: prefix to prepend to each column name
    :return: list of flattened column names
    """
    columns = []
    for field in dataclasses.fields(dataclass_type):
        inner_type = unwrap_optional(field.type)
        name = f"{prefix}{field.name}"
        if dataclasses.is_dataclass(inner_type):
            columns.extend(_flatten_columns(inner_type, prefix=f"{name}."))
        else:
            columns.append(name)
    return columns


def _flatten_instance(instance: object, dataclass_type: type, prefix: str = "") -> dict:
    """
    Flatten a dataclass instance into a dict of column name to formatted value.

    :param instance: the dataclass instance to flatten, or None
    :param dataclass_type: the dataclass type describing `instance`'s shape
    :param prefix: prefix to prepend to each column name
    :return: dict mapping flattened column names to values formatted via
             :func:`format_value`
    """
    row = {}
    for field in dataclasses.fields(dataclass_type):
        inner_type = unwrap_optional(field.type)
        name = f"{prefix}{field.name}"
        value = getattr(instance, field.name, None) if instance is not None else None
        if dataclasses.is_dataclass(inner_type):
            row.update(_flatten_instance(value, inner_type, prefix=f"{name}."))
        else:
            row[name] = format_value(value)
    return row


def _flatten_list_fields(
    instance: object, dataclass_type: type, prefix: str = ""
) -> list[tuple[str, type, list]]:
    """
    Find dot-prefixed columns (matching :func:`_flatten_columns`' naming) whose field
    type is `list[dataclass]`, recursing through nested singular dataclasses the same
    way `_flatten_columns` does so the two stay in sync.

    A `list[dataclass]` field can't be inlined as fixed columns (arbitrary length), so
    instead of JSON-encoding it into one cell, callers use this to spawn a pooled child
    table for `elem_type` and link to it.

    :param instance: the dataclass instance to inspect, or None
    :param dataclass_type: the dataclass type describing `instance`'s shape
    :param prefix: prefix to prepend to each column name
    :return: list of (column_name, element dataclass type, list of item instances)
    """
    results = []
    for field in dataclasses.fields(dataclass_type):
        inner_type = unwrap_optional(field.type)
        name = f"{prefix}{field.name}"
        value = getattr(instance, field.name, None) if instance is not None else None
        if dataclasses.is_dataclass(inner_type):
            results.extend(_flatten_list_fields(value, inner_type, prefix=f"{name}."))
        elif typing.get_origin(inner_type) is list:
            (elem_type,) = typing.get_args(inner_type)
            if dataclasses.is_dataclass(elem_type):
                results.append((name, elem_type, value or []))
    return results


@dataclasses.dataclass(frozen=True)
class _SheetLink:
    """Placeholder cell value meaning "link to this other sheet" rather than raw data."""

    sheet_name: str


@dataclasses.dataclass
class _Table:
    """
    Flattened tabular data for a custom-type attribute or a pooled list[dataclass] field.

    `include_parent_columns` distinguishes root attribute tables (exactly one implicit
    parent - the Summary row, so a parent link adds no information) from pooled child
    tables (reached from a list[dataclass] field at any depth, which can have rows from
    multiple distinct parent rows, so each row needs its own parent link).
    """

    columns: list[str]
    rows: list[list]
    parent_refs: list[tuple[str, int]] = dataclasses.field(default_factory=list)
    include_parent_columns: bool = False


def _process_record(
    instance: object,
    dataclass_type: type,
    sheet_name: str,
    parent_ref: tuple[str, int] | None,
    tables: dict[str, _Table],
    include_parent_columns: bool,
) -> int:
    """
    Flatten `instance` into a new row of `tables[sheet_name]`, creating the table on
    first use, and recurse into any nested list[dataclass] fields (at any depth) to
    populate their own pooled child tables instead of JSON-encoding them into one cell.

    :param instance: the dataclass instance to flatten into a row
    :param dataclass_type: the dataclass type describing `instance`'s shape
    :param sheet_name: name of the table this row belongs to
    :param parent_ref: (parent_sheet_name, parent_row_number) this row is linked from,
                        or None for a root attribute's table
    :param tables: shared dict of sheet name to `_Table`, populated in place
    :param include_parent_columns: whether this row should carry a parent link
    :return: 1-indexed row number of the newly appended row within `tables[sheet_name]`
    """
    columns = _flatten_columns(dataclass_type)
    values = _flatten_instance(instance, dataclass_type)
    list_fields = _flatten_list_fields(instance, dataclass_type)

    table = tables.setdefault(
        sheet_name,
        _Table(columns=columns, rows=[], include_parent_columns=include_parent_columns),
    )
    row_values = [values[column] for column in columns]
    row_number = len(table.rows) + 1

    for column_name, elem_type, sub_items in list_fields:
        col_index = columns.index(column_name)
        if not sub_items:
            row_values[col_index] = NOT_FOUND
            continue
        child_sheet = _sanitize_sheet_name(elem_type.__name__)
        row_values[col_index] = _SheetLink(child_sheet)
        for sub_item in sub_items:
            _process_record(
                sub_item,
                elem_type,
                child_sheet,
                (sheet_name, row_number),
                tables,
                include_parent_columns=True,
            )

    table.rows.append(row_values)
    if include_parent_columns:
        table.parent_refs.append(parent_ref)
    return row_number


def _classify_attribute(
    attr: Attribute, prediction: dspy.Prediction, tables: dict[str, _Table]
) -> object:
    """
    Classify a top-level attribute's value for presentation.

    :param attr: the attribute definition, used for its type
    :param prediction: the DSPy Prediction holding the extracted value
    :param tables: shared dict of sheet name to `_Table`, populated in place for
                    custom-type attributes
    :return: a `_SheetLink` if `attr` is a custom type (or a list of one) with a
             non-empty value, NOT_FOUND if it's a custom type with no value, otherwise
             the value formatted via :func:`format_value`
    """
    value = getattr(prediction, attr.name, None)
    inner_type = unwrap_optional(attr.attr_type)

    dataclass_type = None
    items = []
    if typing.get_origin(inner_type) is list:
        (elem_type,) = typing.get_args(inner_type)
        if dataclasses.is_dataclass(elem_type):
            dataclass_type, items = elem_type, value or []
    elif dataclasses.is_dataclass(inner_type):
        dataclass_type = inner_type
        items = [value] if value is not None else []

    if dataclass_type is None:
        return format_value(value)
    if not items:
        return NOT_FOUND

    sheet_name = _sanitize_sheet_name(attr.name)
    for item in items:
        _process_record(
            item, dataclass_type, sheet_name, None, tables, include_parent_columns=False
        )
    return _SheetLink(sheet_name)


def _append_link_row(
    sheet: openpyxl.worksheet.worksheet.Worksheet, name: str, target_sheet_name: str
) -> None:
    """
    Append a [name, link] row, where the link cell jumps to another sheet.

    :param sheet: the sheet to append the row to
    :param name: value for the first ("name") column
    :param target_sheet_name: title of the sheet the link cell should jump to
    """
    sheet.append([name, target_sheet_name])
    link_cell = sheet.cell(row=sheet.max_row, column=2)
    link_cell.hyperlink = f"#'{target_sheet_name}'!A1"
    link_cell.style = "Hyperlink"


def write_excel(
    prediction: dspy.Prediction, attributes: list[Attribute], path: Path
) -> None:
    """
    Write a prediction's results to an Excel workbook.

    A "Summary" sheet has one row per top-level attribute. Plain
    attributes show their formatted value directly. Attributes that are
    a custom type, or a non-empty list of a custom type, instead get a
    link to a dedicated sheet named after the attribute: one row per
    item, with columns for each field. Fields that are themselves a
    custom type are flattened into dot-prefixed columns (e.g.
    "intervention_type.type_of_intervention"), at any nesting depth.
    Fields that are a *list* of a custom type - at any depth, not just
    top-level attributes - instead link to their own sheet named after
    that type, pooling every instance of it found anywhere in the
    results. Each row in a pooled sheet carries "_parent_sheet" and
    "_parent_row" columns identifying the specific row that produced it,
    since a pooled type can be reached from more than one place.
    Custom-type attributes with no value are shown as NOT_FOUND in the
    Summary sheet, with no dedicated sheet.

    :param prediction: the DSPy Prediction holding the extracted values
    :param attributes: the attribute definitions extracted
    :param path: destination file path
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Summary")
    summary.append(["name", "value"])

    tables: dict[str, _Table] = {}
    for attr in attributes:
        result = _classify_attribute(attr, prediction, tables)
        if isinstance(result, _SheetLink):
            _append_link_row(summary, attr.name, result.sheet_name)
        else:
            summary.append([attr.name, result])

    reasoning = getattr(prediction, "reasoning", None)
    if reasoning is not None:
        summary.append(["_reasoning_", reasoning])

    for sheet_name, table in tables.items():
        sheet = workbook.create_sheet(sheet_name)
        header = (
            ["_parent_sheet", "_parent_row"] + table.columns
            if table.include_parent_columns
            else table.columns
        )
        sheet.append(header)

        parent_refs = (
            table.parent_refs
            if table.include_parent_columns
            else [None] * len(table.rows)
        )
        link_offset = 2 if table.include_parent_columns else 0
        for row_values, parent_ref in zip(table.rows, parent_refs):
            display_values = [
                v.sheet_name if isinstance(v, _SheetLink) else v for v in row_values
            ]
            if table.include_parent_columns:
                display_values = [parent_ref[0], parent_ref[1]] + display_values
            sheet.append(display_values)

            for col_index, value in enumerate(row_values):
                if isinstance(value, _SheetLink):
                    cell = sheet.cell(
                        row=sheet.max_row, column=col_index + 1 + link_offset
                    )
                    cell.hyperlink = f"#'{value.sheet_name}'!A1"
                    cell.style = "Hyperlink"

    workbook.save(path)
